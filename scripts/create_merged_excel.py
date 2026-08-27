import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Dataset Statistics"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
subsection_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
note_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

header_font = Font(bold=True, size=12, color="FFFFFF")
section_font = Font(bold=True, size=11)
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add title
ws.merge_cells('A1:D1')
ws['A1'] = "COMPLETE DATASET STATISTICS - MERGED"
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal='center')
ws['A1'].fill = header_fill
ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")

row = 3

# PRIMARY TRAINING DATASET
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "PRIMARY TRAINING DATASET (ROUND 1 SUBMISSION)"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

# Headers
ws[f'A{row}'] = "Category"
ws[f'B{row}'] = "Metric"
ws[f'C{row}'] = "Value"
ws[f'D{row}'] = "Details"
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].fill = note_fill
row += 1

# Railway Wagon Dataset details
wagon_data = [
    ["Railway Wagon Dataset", "Dataset Name", "Railway Wagon Dataset", "Custom railway wagon video frames"],
    ["Railway Wagon Dataset", "Location", "train/train/", "blur/ and sharp/ folders"],
    ["Railway Wagon Dataset", "Total Image Pairs", "1066 (862 used for training)", "959 training + 107 validation"],
    ["Railway Wagon Dataset", "Training Set (90%)", "959", ""],
    ["Railway Wagon Dataset", "Validation Set (10%)", "107", ""],
    ["Railway Wagon Dataset", "Image Height (pixels)", "506", ""],
    ["Railway Wagon Dataset", "Image Width (pixels)", "898", ""],
    ["Railway Wagon Dataset", "Image Channels", "3", "RGB"],
    ["Railway Wagon Dataset", "Total Pixels per Image", "1,364,876", ""],
    ["Railway Wagon Dataset", "Total Dataset Pixels", "1,454,557,816", ""],
    ["Railway Wagon Dataset", "Format", "JPEG/PNG", "Lossless format for quality preservation"],
    ["Railway Wagon Dataset", "Color Space", "RGB", ""],
    ["Railway Wagon Dataset", "Bit Depth", "8-bit (uint8)", ""],
    ["Railway Wagon Dataset", "Blur Type", "Motion blur", "From moving railway wagons"],
    ["Railway Wagon Dataset", "Ground Truth", "Sharp reference frames", "From same scene"],
    ["Railway Wagon Dataset", "Blur Severity", "91.7% sharpness loss", "Severe motion blur"],
    ["Railway Wagon Dataset", "Average Pixel Difference", "6.39", ""],
    ["Railway Wagon Dataset", "Blur Laplacian Variance", "67.07", ""],
    ["Railway Wagon Dataset", "Sharp Laplacian Variance", "808.92", ""],
    ["Railway Wagon Dataset", "Total Dataset Size", "495.2 MB", "247.6 MB (blur) + 247.6 MB (sharp)"],
    ["Railway Wagon Dataset", "Average Image Size", "0.29 MB", "Per image (approx 287 KB)"],
    ["Railway Wagon Dataset", "Image Resolution", "Variable", "Crops of 256x256 used during training"],
    ["Railway Wagon Dataset", "Preprocessing", "Random Crop", "256x256 random crops, normalization to [0,1]"],
    ["Railway Wagon Dataset", "Augmentation", "Random Cropping + Shuffling", "Different crops each epoch"],
    ["Railway Wagon Dataset", "Training Epochs", "20", "Final fine-tuning on wagon dataset (Stage 3)"],
    ["Railway Wagon Dataset", "Learning Rate", "1e-5", "Fine-tuning learning rate"],
    ["Railway Wagon Dataset", "Batch Size", "2", "Memory-constrained configuration"],
    ["Railway Wagon Dataset", "Purpose", "Final Fine-tuning", "Stage 3 - Adapt to railway wagon specific patterns"],
    ["Railway Wagon Dataset", "Submitted Model Performance", "29.86 dB", "After multi-stage training pipeline"],
]

for data in wagon_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# PRE-TRAINING DATASETS
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "PRE-TRAINING DATASETS (MULTI-STAGE APPROACH)"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

# GoPro Dataset
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "1. GoPro Dataset"
ws[f'A{row}'].fill = subsection_fill
ws[f'A{row}'].font = bold_font
row += 1

gopro_data = [
    ["1. GoPro Dataset", "Total Image Pairs", "2,185", "GoPro motion blur dataset"],
    ["1. GoPro Dataset", "Training Images (Blur)", "2,185", "Motion-blurred images"],
    ["1. GoPro Dataset", "Training Images (Sharp)", "2,185", "Ground truth sharp images"],
    ["1. GoPro Dataset", "Domain", "Motion Deblurring", "Generic camera motion blur"],
    ["1. GoPro Dataset", "Image Format", "PNG", "High-quality image format"],
    ["1. GoPro Dataset", "Training Epochs", "70", "Full training from scratch"],
    ["1. GoPro Dataset", "Learning Rate", "2e-4", "Cosine annealing scheduler"],
    ["1. GoPro Dataset", "Batch Size", "4", "For pre-training"],
    ["1. GoPro Dataset", "Purpose", "Pre-training", "Learn generic deblurring features"],
    ["1. GoPro Dataset", "Location", "../GOPRO_Large/", "Pre-training dataset"],
    ["1. GoPro Dataset", "Status", "Used for pre-training", "Stage 1 of training pipeline"],
]

for data in gopro_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# LOL Dataset
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "2. LOL Blur Dataset"
ws[f'A{row}'].fill = subsection_fill
ws[f'A{row}'].font = bold_font
row += 1

lol_data = [
    ["2. LOL Blur Dataset", "Total Image Pairs", "489", "Low-light + blur dataset"],
    ["2. LOL Blur Dataset", "Training Images (Blur)", "489", "Low-light blurred images"],
    ["2. LOL Blur Dataset", "Training Images (Sharp)", "489", "Enhanced sharp images"],
    ["2. LOL Blur Dataset", "Domain", "Low-Light + Blur", "Combined low-light and motion blur"],
    ["2. LOL Blur Dataset", "Image Format", "PNG", "High-quality image format"],
    ["2. LOL Blur Dataset", "Training Epochs", "30", "Fine-tuning on GoPro weights"],
    ["2. LOL Blur Dataset", "Learning Rate", "1e-5", "Fine-tuning learning rate"],
    ["2. LOL Blur Dataset", "Batch Size", "2", "For fine-tuning"],
    ["2. LOL Blur Dataset", "Purpose", "Fine-tuning", "Adapt to low-light conditions"],
    ["2. LOL Blur Dataset", "Location", "../LOL_BLUR/", "Fine-tuning dataset"],
    ["2. LOL Blur Dataset", "Status", "Used for continued training", "Stage 2 of training pipeline"],
]

for data in lol_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# TRAINING SUMMARY
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "TRAINING SUMMARY"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

summary_data = [
    ["Total Training Images", "Total Images Used", "3,536", "2,185 (GoPro) + 489 (LOL) + 862 (Wagon)"],
    ["Total Training Images", "Round 1 Submission Training", "All datasets used", "Multi-stage training approach"],
    ["Total Training Images", "Training Strategy", "Multi-stage Fine-tuning", "GoPro → LOL → Wagon (sequential)"],
    ["Total Training Images", "Batch Size", "2-4", "2 for wagon/LOL fine-tuning; 4 for GoPro"],
    ["Total Training Images", "Memory Configuration", "Memory-constrained", "GPU memory optimization"],
]

for data in summary_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# TRAINING PIPELINE STAGES
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "TRAINING PIPELINE STAGES"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

stages_data = [
    ["Stage 1", "Pre-training on GoPro", "2,185 images", "Generic motion deblurring features"],
    ["Stage 2", "Fine-tuning on LOL Blur", "489 images", "Low-light adaptation"],
    ["Stage 3", "Final fine-tuning on Wagon", "862 images", "Railway wagon domain specialization"],
    ["Final Performance", "Validation PSNR", "29.86 dB", "On railway wagon validation set"],
]

for data in stages_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# IMPORTANT CLARIFICATION
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "IMPORTANT CLARIFICATION"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

notes_data = [
    ["Note 1", "Training Approach", "Multi-stage Training", "Model trained using GoPro → LOL → Wagon pipeline"],
    ["Note 2", "Dataset Utilization", "All Datasets Used", "GoPro, LOL, and Railway Wagon datasets all used"],
    ["Note 3", "Training Strategy", "Transfer Learning", "Sequential fine-tuning from generic to domain-specific"],
    ["Note 4", "Final Specialization", "Railway Wagon Focus", "Final fine-tuning ensures domain-specific performance"],
    ["Note 5", "Validation Performance", "29.86 dB PSNR", "Achieved on railway wagon validation set"],
]

for data in notes_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

# Auto-adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 50

# Apply borders to all cells
for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row_cells:
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Save the workbook
output_path = r'HACKATHON_SUBMISSION\MERGED_DATASET_STATISTICS.xlsx'
wb.save(output_path)
print(f"Excel file created successfully: {output_path}")
