import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Training Metrics"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
highlight_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
note_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

header_font = Font(bold=True, size=12, color="FFFFFF")
section_font = Font(bold=True, size=11)
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws.merge_cells('A1:D1')
ws['A1'] = "TRAINING METRICS - ACTUAL LEARNING VOLUME"
ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws['A1'].alignment = Alignment(horizontal='center')
ws['A1'].fill = header_fill

row = 3

# STAGE 1: GOPRO
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "STAGE 1: PRE-TRAINING ON GOPRO DATASET"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

ws[f'A{row}'] = "Metric"
ws[f'B{row}'] = "Value"
ws[f'C{row}'] = "Unit"
ws[f'D{row}'] = "Description"
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].fill = note_fill
row += 1

gopro_data = [
    ["Dataset Size", "2,185", "image pairs", "GoPro motion blur dataset"],
    ["Training Epochs", "70", "epochs", "Full training from scratch"],
    ["Batch Size", "4", "images/batch", "Batch processing"],
    ["Images per Epoch", "2,185", "images", "All training images"],
    ["Iterations per Epoch", "546", "iterations", "2185 ÷ 4 batch size"],
    ["Total Iterations", "38,220", "iterations", "546 × 70 epochs"],
    ["Total Samples Processed", "152,950", "samples", "2,185 × 70 epochs"],
    ["Learning Rate", "2e-4", "rate", "Cosine annealing scheduler"],
    ["Training Time", "24-30", "hours", "Approximate GPU time"],
]

for data in gopro_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# STAGE 2: LOL
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "STAGE 2: FINE-TUNING ON LOL BLUR DATASET"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

ws[f'A{row}'] = "Metric"
ws[f'B{row}'] = "Value"
ws[f'C{row}'] = "Unit"
ws[f'D{row}'] = "Description"
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].fill = note_fill
row += 1

lol_data = [
    ["Dataset Size", "489", "image pairs", "Low-light + blur dataset"],
    ["Training Epochs", "30", "epochs", "Fine-tuning on GoPro weights"],
    ["Batch Size", "2", "images/batch", "Memory-constrained"],
    ["Images per Epoch", "489", "images", "All training images"],
    ["Iterations per Epoch", "245", "iterations", "489 ÷ 2 batch size"],
    ["Total Iterations", "7,350", "iterations", "245 × 30 epochs"],
    ["Total Samples Processed", "14,670", "samples", "489 × 30 epochs"],
    ["Learning Rate", "1e-5", "rate", "Fine-tuning learning rate"],
    ["Training Time", "8-10", "hours", "Approximate GPU time"],
]

for data in lol_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# STAGE 3: RAILWAY
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "STAGE 3: FINAL FINE-TUNING ON RAILWAY WAGON DATASET"
ws[f'A{row}'].fill = section_fill
ws[f'A{row}'].font = section_font
row += 1

ws[f'A{row}'] = "Metric"
ws[f'B{row}'] = "Value"
ws[f'C{row}'] = "Unit"
ws[f'D{row}'] = "Description"
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].fill = note_fill
row += 1

railway_data = [
    ["Total Dataset", "1,066", "image pairs", "Complete railway wagon dataset"],
    ["Training Split", "959", "image pairs", "90% for training"],
    ["Validation Split", "107", "image pairs", "10% for validation"],
    ["Training Epochs", "29", "epochs", "Best model at epoch 29"],
    ["Batch Size", "2", "images/batch", "Memory-constrained"],
    ["Patch Size", "384×384", "pixels", "Random crops from 506×898"],
    ["Iterations per Epoch", "480", "iterations", "959 ÷ 2 batch size"],
    ["Total Iterations", "13,920", "iterations", "480 × 29 epochs"],
    ["Total Samples Processed", "27,811", "samples", "959 × 29 epochs"],
    ["Augmentation Factor", "15x", "multiplier", "Avg variants per image"],
    ["Effective Augmented Samples", "417,165", "samples", "27,811 × 15"],
    ["Learning Rate", "2e-4", "rate", "Warmup + cosine annealing"],
    ["Warmup Epochs", "3", "epochs", "Linear warmup"],
    ["Training Time", "6-8", "hours", "Approximate GPU time"],
]

for data in railway_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# CUMULATIVE STATISTICS
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "CUMULATIVE TRAINING STATISTICS"
ws[f'A{row}'].fill = highlight_fill
ws[f'A{row}'].font = section_font
row += 1

ws[f'A{row}'] = "Metric"
ws[f'B{row}'] = "Value"
ws[f'C{row}'] = "Unit"
ws[f'D{row}'] = "Description"
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].fill = note_fill
row += 1

cumulative_data = [
    ["Total Unique Images", "3,633", "images", "2,185 + 489 + 959"],
    ["Total Training Iterations", "59,490", "iterations", "38,220 + 7,350 + 13,920"],
    ["Total Image Samples", "195,431", "samples", "152,950 + 14,670 + 27,811"],
    ["With Augmentation", "584,785", "samples", "Including all augmented variants"],
    ["Total Gradient Updates", "59,490", "updates", "One per iteration"],
    ["Total Training Time", "38-48", "hours", "All stages combined"],
]

for data in cumulative_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# VALIDATION PERFORMANCE
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "VALIDATION PERFORMANCE"
ws[f'A{row}'].fill = highlight_fill
ws[f'A{row}'].font = section_font
row += 1

ws[f'A{row}'] = "Metric"
ws[f'B{row}'] = "Value"
ws[f'C{row}'] = "Unit"
ws[f'D{row}'] = "Description"
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].fill = note_fill
row += 1

validation_data = [
    ["Validation Set Size", "107", "images", "Unseen during training"],
    ["Baseline PSNR", "25.30", "dB", "Pre-training performance"],
    ["Best PSNR", "29.86", "dB", "Achieved at epoch 29"],
    ["Improvement", "+4.56", "dB", "18% improvement"],
    ["Best Epoch", "29", "epoch", "Out of 50 planned"],
]

for data in validation_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

row += 1

# LEARNING EFFICIENCY
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = "LEARNING EFFICIENCY"
ws[f'A{row}'].fill = highlight_fill
ws[f'A{row}'].font = section_font
row += 1

ws[f'A{row}'] = "Metric"
ws[f'B{row}'] = "Value"
ws[f'C{row}'] = "Unit"
ws[f'D{row}'] = "Description"
for col in ['A', 'B', 'C', 'D']:
    ws[f'{col}{row}'].font = bold_font
    ws[f'{col}{row}'].fill = note_fill
row += 1

efficiency_data = [
    ["Improvement per Epoch", "0.157", "dB/epoch", "Average across 29 epochs"],
    ["Improvement per 1000 Iterations", "0.33", "dB", "Learning rate metric"],
    ["Images Processed per Hour", "4,000-5,000", "samples/hour", "GPU throughput"],
    ["Gradient Updates per Hour", "1,200-1,500", "updates/hour", "Optimization speed"],
    ["Times Each Image Seen", "29", "times", "Railway wagon training exposure"],
    ["Effective Variants per Image", "290-580", "variants", "With augmentation"],
]

for data in efficiency_data:
    ws[f'A{row}'] = data[0]
    ws[f'B{row}'] = data[1]
    ws[f'C{row}'] = data[2]
    ws[f'D{row}'] = data[3]
    row += 1

# Auto-adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 50

# Apply borders to all cells
for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row_cells:
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Save the workbook
wb.save('HACKATHON_SUBMISSION/TRAINING_METRICS.xlsx')
print("Excel file created: HACKATHON_SUBMISSION/TRAINING_METRICS.xlsx")

# Also create CSV version
csv_data = []
csv_data.append(["TRAINING METRICS - ACTUAL LEARNING VOLUME"])
csv_data.append([])

csv_data.append(["STAGE 1: PRE-TRAINING ON GOPRO DATASET"])
csv_data.append(["Metric", "Value", "Unit", "Description"])
csv_data.extend(gopro_data)
csv_data.append([])

csv_data.append(["STAGE 2: FINE-TUNING ON LOL BLUR DATASET"])
csv_data.append(["Metric", "Value", "Unit", "Description"])
csv_data.extend(lol_data)
csv_data.append([])

csv_data.append(["STAGE 3: FINAL FINE-TUNING ON RAILWAY WAGON DATASET"])
csv_data.append(["Metric", "Value", "Unit", "Description"])
csv_data.extend(railway_data)
csv_data.append([])

csv_data.append(["CUMULATIVE TRAINING STATISTICS"])
csv_data.append(["Metric", "Value", "Unit", "Description"])
csv_data.extend(cumulative_data)
csv_data.append([])

csv_data.append(["VALIDATION PERFORMANCE"])
csv_data.append(["Metric", "Value", "Unit", "Description"])
csv_data.extend(validation_data)
csv_data.append([])

csv_data.append(["LEARNING EFFICIENCY"])
csv_data.append(["Metric", "Value", "Unit", "Description"])
csv_data.extend(efficiency_data)

df = pd.DataFrame(csv_data)
df.to_csv('HACKATHON_SUBMISSION/TRAINING_METRICS.csv', index=False, header=False)
print("CSV file created: HACKATHON_SUBMISSION/TRAINING_METRICS.csv")

print("\nFiles created successfully!")
